from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────
GOLD   = "B8975A"
GOLD_L = "D4AF6E"
DARK   = "1A1A18"
DARK2  = "2A2A28"
OFF_W  = "F5F0E8"
WHITE  = "FFFFFF"
GREEN  = "4CAF7D"
RED    = "E05454"
BLUE   = "5B8DF5"
GREY   = "3A3A38"
LT_GLD = "F5EDD8"   # light gold fill for alt rows

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color=WHITE, bold=False, size=11, name="Calibri"):
    return Font(color=hex_color, bold=bold, size=size, name=name)

def border(style="thin"):
    s = Side(style=style, color="2A2A28")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════
# SHEET 1 — COVER
# ═══════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 4
ws_cover.column_dimensions["B"].width = 26
ws_cover.column_dimensions["C"].width = 26
ws_cover.column_dimensions["D"].width = 20
ws_cover.column_dimensions["E"].width = 20
ws_cover.column_dimensions["F"].width = 4

# Dark background for whole cover
for r in range(1, 50):
    for c in range(1, 7):
        ws_cover.cell(r, c).fill = fill(DARK)

# Title block
ws_cover.row_dimensions[3].height = 14
ws_cover.row_dimensions[4].height = 44
ws_cover.row_dimensions[5].height = 24
ws_cover.row_dimensions[6].height = 18
ws_cover.row_dimensions[7].height = 14
ws_cover.row_dimensions[8].height = 18

eyebrow = ws_cover.cell(3, 2, "MERCHANT CLUB SA")
eyebrow.font  = Font(color=GOLD, bold=True, size=9, name="Calibri")
eyebrow.alignment = left()

title = ws_cover.cell(4, 2, "User Acceptance Testing")
ws_cover.merge_cells("B4:E4")
title.font      = Font(color=WHITE, bold=True, size=28, name="Calibri")
title.alignment = left()

sub = ws_cover.cell(5, 2, "Platform Build — Phase 1 through 5")
ws_cover.merge_cells("B5:E5")
sub.font      = Font(color=GOLD_L, size=13, name="Calibri")
sub.alignment = left()

# Gold rule
for c in range(2, 6):
    cell = ws_cover.cell(6, c)
    cell.fill = fill(GOLD)
    cell.value = ""

# Meta info
meta = [
    ("Version",   "1.0"),
    ("Date",      "April 2026"),
    ("Site",      "merchantclubsa.com"),
    ("Prepared",  "Nexus Team"),
    ("Platform",  "Next.js 16 · Supabase · Vercel · Resend"),
]
for i, (lbl, val) in enumerate(meta):
    row = 8 + i
    ws_cover.row_dimensions[row].height = 22
    lc = ws_cover.cell(row, 2, lbl)
    lc.font      = Font(color=GOLD_L, size=10, name="Calibri")
    lc.alignment = left()
    vc = ws_cover.cell(row, 3, val)
    ws_cover.merge_cells(f"C{row}:E{row}")
    vc.font      = Font(color=OFF_W, size=10, name="Calibri")
    vc.alignment = left()

# Test Summary Table
ws_cover.row_dimensions[15].height = 10
ws_cover.row_dimensions[16].height = 26
ws_cover.row_dimensions[17].height = 20
ws_cover.row_dimensions[18].height = 20
ws_cover.row_dimensions[19].height = 20
ws_cover.row_dimensions[20].height = 20
ws_cover.row_dimensions[21].height = 20
ws_cover.row_dimensions[22].height = 20
ws_cover.row_dimensions[23].height = 20
ws_cover.row_dimensions[24].height = 20

# Summary header
ws_cover.merge_cells("B16:E16")
sh = ws_cover.cell(16, 2, "Test Summary")
sh.fill      = fill(DARK2)
sh.font      = Font(color=GOLD, bold=True, size=12, name="Calibri")
sh.alignment = center()
sh.border    = Border(bottom=Side(style="medium", color=GOLD))

col_hdrs = ["Area", "Total Tests", "Pass", "Fail"]
for ci, hdr in enumerate(col_hdrs, 2):
    c = ws_cover.cell(17, ci, hdr)
    c.fill      = fill(GOLD)
    c.font      = Font(color=DARK, bold=True, size=10, name="Calibri")
    c.alignment = center()
    c.border    = border()

areas = [
    ("1 — Storefront",          6),
    ("2 — Customer Checkout",   7),
    ("3 — Brand Order Mgmt",    6),
    ("4 — Email Notifications", 5),
    ("5 — Creator Members",     5),
    ("6 — Admin Dashboard",     8),
    ("7 — Stock Management",    5),
]
for i, (area, total) in enumerate(areas):
    row = 18 + i
    bg = DARK2 if i % 2 == 0 else DARK
    cells_data = [area, total, "—", "—"]
    for ci, val in enumerate(cells_data, 2):
        c = ws_cover.cell(row, ci, val)
        c.fill      = fill(bg)
        c.font      = Font(color=OFF_W if ci > 2 else GOLD_L, size=10)
        c.alignment = center() if ci > 2 else left()
        c.border    = border()

total_row = 26
ws_cover.row_dimensions[total_row].height = 24
totals = ["TOTAL", 42, "—", "—"]
for ci, val in enumerate(totals, 2):
    c = ws_cover.cell(total_row, ci, val)
    c.fill      = fill(GOLD)
    c.font      = Font(color=DARK, bold=True, size=10)
    c.alignment = center() if ci > 2 else left()
    c.border    = border()

ws_cover.row_dimensions[28].height = 18
note = ws_cover.cell(28, 2,
    "Status: PASS / FAIL / BLOCKED / N/A — Fill in UAT Tests sheet")
ws_cover.merge_cells("B28:E28")
note.font      = Font(color=GREY, size=9, italic=True)
note.alignment = left()

# ═══════════════════════════════════════════
# SHEET 2 — UAT TESTS
# ═══════════════════════════════════════════
ws_uat = wb.create_sheet("UAT Tests")
ws_uat.sheet_view.showGridLines = False

col_widths = [8, 24, 36, 44, 36, 14, 24, 18, 14]
col_names  = ["ID", "Area", "Test Case", "Steps",
              "Expected Result", "Status", "Notes", "Tested By", "Date"]
for i, (w, h) in enumerate(zip(col_widths, col_names), 1):
    ws_uat.column_dimensions[get_column_letter(i)].width = w

ws_uat.row_dimensions[1].height = 36
for ci, hdr in enumerate(col_names, 1):
    c = ws_uat.cell(1, ci, hdr)
    c.fill      = fill(DARK)
    c.font      = Font(color=GOLD, bold=True, size=10, name="Calibri")
    c.alignment = center()
    c.border    = border("medium")

ws_uat.freeze_panes = "A2"

# Data validation for Status column (F)
dv = DataValidation(
    type="list",
    formula1='"PASS,FAIL,BLOCKED,N/A"',
    allow_blank=True,
    showDropDown=False,
)
dv.error       = "Use: PASS, FAIL, BLOCKED, or N/A"
dv.errorTitle  = "Invalid Status"
dv.prompt      = "Select test status"
dv.promptTitle = "Status"
dv.sqref       = "F2:F200"
ws_uat.add_data_validation(dv)

tests = [
    # ID, Area, Test Case, Steps, Expected Result
    ("TC-001","1 — Storefront","Brands page shows live brands from DB",
     "1. Navigate to /en/brands\n2. Observe main content section",
     "Active Partners section visible with real brand cards pulled from database (not Coming Soon placeholder)"),

    ("TC-002","1 — Storefront","Brand page loads correctly",
     "1. Click on Test01 brand card\n2. Observe page content",
     "Brand detail page loads with logo, name, description, and product listings"),

    ("TC-003","1 — Storefront","Product page loads",
     "1. Navigate to /en/brands/test01\n2. Click any product\n3. Observe product page",
     "Product page shows title, images, price, stock level, and Order button"),

    ("TC-004","1 — Storefront","AR/EN language switch works",
     "1. Navigate to /ar/brands\n2. Verify Arabic layout\n3. Switch back to /en",
     "All content displays in Arabic; RTL layout applied; switching to EN restores English content"),

    ("TC-005","1 — Storefront","Coming soon shows when no live brands",
     "1. Set all products to non-live status in Supabase\n2. Visit /en/brands",
     "Coming Soon placeholder grid displays instead of active partner section"),

    ("TC-006","1 — Storefront","Out-of-stock product cannot be ordered",
     "1. Find product with stock_quantity = 0\n2. Attempt to navigate to /order",
     "Redirected away from order page OR error displayed; cannot submit order form"),

    ("TC-007","2 — Customer Checkout","Order placed without login",
     "1. Open product page (not logged in)\n2. Click Order\n3. Fill name, phone, city, address\n4. Submit",
     "Order confirmation page shown with order number; no login was required"),

    ("TC-008","2 — Customer Checkout","Required fields validated",
     "1. Open order form\n2. Leave name blank\n3. Submit\n4. Repeat for phone, city, address",
     "Error message 'Please fill in all required fields' shown; order not submitted"),

    ("TC-009","2 — Customer Checkout","Phone field validation",
     "1. Enter phone with fewer than 9 digits\n2. Submit form",
     "Error 'Please enter a valid phone number' displayed"),

    ("TC-010","2 — Customer Checkout","Confirmation page shows correct details",
     "1. Complete a full order\n2. Observe confirmation page",
     "Order number, product name, price, and delivery address all match what was entered"),

    ("TC-011","2 — Customer Checkout","Order saved in Supabase",
     "1. Place a test order\n2. Check Supabase orders table",
     "New row appears in orders table with correct brand_id, customer data, items JSON, subtotal, status=pending"),

    ("TC-012","2 — Customer Checkout","Quantity field works",
     "1. Change quantity to 2 on order form\n2. Submit",
     "Subtotal = unit price × 2; items array in Supabase shows quantity: 2"),

    ("TC-013","2 — Customer Checkout","Order blocked when qty > stock",
     "1. Find product with stock_quantity = 3\n2. Enter quantity = 5\n3. Submit",
     "Error 'Only 3 units available' shown; order not created in DB"),

    ("TC-014","3 — Brand Order Mgmt","Brand sees their orders",
     "1. Log in as brand user\n2. Navigate to /dashboard/brand/orders",
     "Orders table displays with stat boxes (Total, Pending, Shipped, Revenue) and correct filter tabs"),

    ("TC-015","3 — Brand Order Mgmt","Brand confirms pending order",
     "1. Click on a pending order row to expand\n2. Click 'Confirm Order'",
     "Status changes to Confirmed; pill updates to blue; revalidated without page reload"),

    ("TC-016","3 — Brand Order Mgmt","Brand marks as shipped with tracking",
     "1. Expand confirmed order\n2. Enter tracking number\n3. Click 'Mark as Shipped'",
     "Status = Shipped; tracking number saved in DB and displayed on row"),

    ("TC-017","3 — Brand Order Mgmt","Brand marks as delivered",
     "1. Expand shipped order\n2. Click 'Mark as Delivered'",
     "Status = Delivered; green pill shown; no further action buttons available"),

    ("TC-018","3 — Brand Order Mgmt","Brand cancels order",
     "1. Expand pending or confirmed order\n2. Click 'Cancel Order'",
     "Status = Cancelled; row shows grey pill; no action buttons visible"),

    ("TC-019","3 — Brand Order Mgmt","Filter tabs show correct counts",
     "1. Create orders in various statuses\n2. Click each filter tab on brand orders page",
     "Each tab (Pending/Confirmed/Shipped/Delivered/Cancelled) shows accurate count and filters correctly"),

    ("TC-020","4 — Email Notifications","Customer receives order confirmation",
     "1. Place order with valid customer email address\n2. Check that email inbox",
     "Email received with subject 'Order Confirmed — #MCO-...' containing order number, product, total, delivery address"),

    ("TC-021","4 — Email Notifications","Brand receives new order notification",
     "1. Place any order\n2. Check brand's contact_email inbox",
     "Email with subject 'New Order — #MCO-...' received; contains customer name, phone, city, product, total, and dashboard link"),

    ("TC-022","4 — Email Notifications","Customer gets confirmed status email",
     "1. Brand confirms an order where customer has email on file\n2. Check customer inbox",
     "Email 'Order Confirmed — #...' received with gold accent; body says order is being prepared"),

    ("TC-023","4 — Email Notifications","Customer gets shipped email with tracking",
     "1. Brand marks order as shipped with tracking number\n2. Check customer inbox",
     "Email 'Your Order Has Shipped — #...' received; tracking number displayed in monospace font"),

    ("TC-024","4 — Email Notifications","Customer gets delivered email",
     "1. Brand marks order as delivered\n2. Check customer inbox",
     "Email 'Order Delivered — #...' received with green accent; body says order has been delivered"),

    ("TC-025","5 — Creator Members","Creator submits apply/member form",
     "1. Navigate to /en/apply/member\n2. Fill all required fields\n3. Submit",
     "Success message shown; email notification sent to info@merchantclubsa.com"),

    ("TC-026","5 — Creator Members","Logged-in user gets DB record",
     "1. Log in as any user account\n2. Navigate to /en/apply/member\n3. Submit form",
     "Row inserted in members table with user_id, full_name, email, status='pending'"),

    ("TC-027","5 — Creator Members","Admin approves member",
     "1. Log in as admin\n2. Go to /dashboard/admin/members\n3. Find pending member\n4. Click Approve",
     "Member status changes to Approved; green pill displayed; reviewed_at timestamp set in DB"),

    ("TC-028","5 — Creator Members","Admin rejects member",
     "1. Log in as admin\n2. Go to /dashboard/admin/members\n3. Find pending member\n4. Click Reject",
     "Member status changes to Rejected; red pill displayed"),

    ("TC-029","5 — Creator Members","Admin revokes approved member",
     "1. Find approved member\n2. Click Revoke",
     "Member status changes to Rejected; row reflects new status immediately"),

    ("TC-030","6 — Admin Dashboard","Overview page loads with stats",
     "1. Log in as admin\n2. Navigate to /dashboard/admin",
     "Overview displays with stat cards, recent activity, and all sidebar navigation items visible"),

    ("TC-031","6 — Admin Dashboard","Brands table — colored status pills",
     "1. Go to /dashboard/admin/brands\n2. Observe status column",
     "Approved = green pill, Pending = gold pill, Suspended = red pill; mini stat boxes at top"),

    ("TC-032","6 — Admin Dashboard","Applications — approve/reject works",
     "1. Go to /dashboard/admin/applications\n2. Find pending application\n3. Click Approve",
     "Brand account created in DB; approval email sent; application status = approved"),

    ("TC-033","6 — Admin Dashboard","Products table — thumbnails and actions",
     "1. Go to /dashboard/admin/products\n2. Observe table",
     "Product thumbnails (40×40) visible; Approve/Reject/Unpublish buttons functional; expand shows image gallery"),

    ("TC-034","6 — Admin Dashboard","Members page loads",
     "1. Go to /dashboard/admin/members",
     "Members table loads with stat boxes (Total, Pending, Approved, Rejected); filter tabs work (requires members table in DB)"),

    ("TC-035","6 — Admin Dashboard","Global search finds brands and products",
     "1. Type 2+ characters in admin topbar search box",
     "Dropdown appears with matching brands, products, orders; clicking a result navigates to correct page"),

    ("TC-036","6 — Admin Dashboard","Light mode readable on all admin pages",
     "1. Toggle light mode via sun icon in admin topbar\n2. Visit each admin page",
     "All admin pages fully readable in light mode; no white-on-white or invisible text"),

    ("TC-037","6 — Admin Dashboard","Admin can change brand status",
     "1. Go to admin/brands\n2. Find approved brand\n3. Click Suspend",
     "Brand status updates to Suspended in DB; red pill displayed; brand removed from public storefront"),

    ("TC-038","7 — Stock Management","Stock decreases after order placed",
     "1. Note stock_quantity for a product in Supabase\n2. Place order for qty=1\n3. Re-check stock_quantity",
     "stock_quantity has decreased by exactly 1"),

    ("TC-039","7 — Stock Management","Product goes out_of_stock when stock hits zero",
     "1. Find product with stock_quantity=1\n2. Place order for qty=1\n3. Check product status in Supabase",
     "product.status changed to 'out_of_stock' in DB automatically"),

    ("TC-040","7 — Stock Management","Cannot order out-of-stock product",
     "1. Set product status to 'out_of_stock' in Supabase\n2. Attempt to navigate to /order page",
     "Product query returns null (filtered by status=live); user redirected to brand page"),

    ("TC-041","7 — Stock Management","Stock never goes negative",
     "1. Find product with stock=1\n2. Try to order qty=2",
     "Error 'Only 1 unit available' shown; stock_quantity remains at 1; no order created"),

    ("TC-042","7 — Stock Management","Admin products table reflects current stock",
     "1. After placing orders, go to /dashboard/admin/products\n2. Check stock display",
     "Correct stock_quantity shown per product; out_of_stock products show status pill correctly"),
]

STATUS_FILLS = {
    "PASS":    "C6EFCE",
    "FAIL":    "FFC7CE",
    "BLOCKED": "FFEB9C",
    "N/A":     "D9D9D9",
}

for ri, (tid, area, case, steps, expected) in enumerate(tests, 2):
    is_even = ri % 2 == 0
    row_bg  = DARK if is_even else DARK2
    ws_uat.row_dimensions[ri].height = 72

    vals = [tid, area, case, steps, expected, "", "", "", ""]
    for ci, val in enumerate(vals, 1):
        c = ws_uat.cell(ri, ci, val)
        c.fill      = fill(row_bg)
        c.font      = Font(color=OFF_W, size=9)
        c.alignment = left()
        c.border    = border()
        if ci == 1:  # ID
            c.font      = Font(color=GOLD, bold=True, size=9)
            c.alignment = center()
        if ci == 2:  # Area
            c.font = Font(color=GOLD_L, size=9)
        if ci == 6:  # Status — leave blank for testers
            c.fill      = fill("2E2E2C")
            c.alignment = center()

# Conditional formatting for Status column
for status, bg in STATUS_FILLS.items():
    ws_uat.conditional_formatting.add(
        f"F2:F{len(tests)+1}",
        FormulaRule(
            formula=[f'F2="{status}"'],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(color=DARK, bold=True),
        )
    )

# ═══════════════════════════════════════════
# SHEET 3 — ISSUES LOG
# ═══════════════════════════════════════════
ws_issues = wb.create_sheet("Issues Log")
ws_issues.sheet_view.showGridLines = False

issue_cols  = ["Issue ID", "Linked TC", "Title", "Steps to Reproduce",
               "Expected", "Actual", "Severity", "Status", "Assigned To", "Resolved"]
issue_widths = [12, 12, 32, 44, 32, 32, 14, 14, 18, 14]

for i, (w, h) in enumerate(zip(issue_widths, issue_cols), 1):
    ws_issues.column_dimensions[get_column_letter(i)].width = w

ws_issues.row_dimensions[1].height = 36
for ci, hdr in enumerate(issue_cols, 1):
    c = ws_issues.cell(1, ci, hdr)
    c.fill      = fill(DARK)
    c.font      = Font(color=GOLD, bold=True, size=10)
    c.alignment = center()
    c.border    = border("medium")

ws_issues.freeze_panes = "A2"

sev_dv = DataValidation(
    type="list",
    formula1='"Critical,High,Medium,Low"',
    allow_blank=True,
)
sev_dv.sqref = "G2:G200"
ws_issues.add_data_validation(sev_dv)

stat_dv = DataValidation(
    type="list",
    formula1='"Open,In Progress,Fixed,Closed,Won\'t Fix"',
    allow_blank=True,
)
stat_dv.sqref = "H2:H200"
ws_issues.add_data_validation(stat_dv)

# Pre-fill known issue
known = ["ISS-001", "TC-039", "T-Shart B stock=0 but status=live",
         "Check T-Shart B product in Supabase admin",
         "status = 'out_of_stock'",
         "status = 'live' (pre-Phase-5 data — not flipped by new logic)",
         "Low", "Open", "Admin", ""]
for ci, val in enumerate(known, 1):
    c = ws_issues.cell(2, ci, val)
    c.fill      = fill(DARK2)
    c.font      = Font(color=OFF_W if ci not in [1, 7, 8] else GOLD_L, size=9)
    c.alignment = left() if ci not in [1, 7, 8] else center()
    c.border    = border()
ws_issues.row_dimensions[2].height = 50

# Empty rows
for ri in range(3, 30):
    ws_issues.row_dimensions[ri].height = 30
    for ci in range(1, 11):
        c = ws_issues.cell(ri, ci)
        c.fill   = fill(DARK if ri % 2 == 0 else DARK2)
        c.border = border()

# ═══════════════════════════════════════════
# SHEET 4 — SUMMARY
# ═══════════════════════════════════════════
ws_sum = wb.create_sheet("Summary")
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 4
ws_sum.column_dimensions["B"].width = 28
ws_sum.column_dimensions["C"].width = 16
ws_sum.column_dimensions["D"].width = 16
ws_sum.column_dimensions["E"].width = 16
ws_sum.column_dimensions["F"].width = 16
ws_sum.column_dimensions["G"].width = 16
ws_sum.column_dimensions["H"].width = 4

for r in range(1, 60):
    for c in range(1, 9):
        ws_sum.cell(r, c).fill = fill(DARK)

# Title
ws_sum.row_dimensions[3].height = 36
ws_sum.merge_cells("B3:G3")
t = ws_sum.cell(3, 2, "UAT Summary Report")
t.font      = Font(color=WHITE, bold=True, size=20)
t.alignment = center()

ws_sum.row_dimensions[4].height = 6
for c in range(2, 8):
    ws_sum.cell(4, c).fill = fill(GOLD)

# Instructions
ws_sum.row_dimensions[5].height = 20
ws_sum.merge_cells("B5:G5")
inst = ws_sum.cell(5, 2,
    "This sheet auto-calculates from the UAT Tests sheet. Fill in Status column there first.")
inst.font      = Font(color=GOLD_L, size=9, italic=True)
inst.alignment = center()

# Area breakdown table header
ws_sum.row_dimensions[7].height = 26
hdr_data = ["Area", "Total", "Pass", "Fail", "Blocked", "Pass Rate"]
hdr_cols  = ["B", "C", "D", "E", "F", "G"]
for col, hdr in zip(hdr_cols, hdr_data):
    c = ws_sum[f"{col}7"]
    c.value     = hdr
    c.fill      = fill(GOLD)
    c.font      = Font(color=DARK, bold=True, size=10)
    c.alignment = center()
    c.border    = border()

# Area rows with COUNTIF formulas
area_labels = [
    ("1 — Storefront",          "TC-001","TC-006"),
    ("2 — Customer Checkout",   "TC-007","TC-013"),
    ("3 — Brand Order Mgmt",    "TC-014","TC-019"),
    ("4 — Email Notifications", "TC-020","TC-024"),
    ("5 — Creator Members",     "TC-025","TC-029"),
    ("6 — Admin Dashboard",     "TC-030","TC-037"),
    ("7 — Stock Management",    "TC-038","TC-042"),
]

for i, (label, start_id, end_id) in enumerate(area_labels):
    row = 8 + i
    ws_sum.row_dimensions[row].height = 22
    bg = DARK2 if i % 2 == 0 else DARK

    # Calculate row range in UAT Tests sheet
    start_row = tests.index(next(t for t in tests if t[0] == start_id)) + 2
    end_row   = tests.index(next(t for t in tests if t[0] == end_id))   + 2
    rng = f"'UAT Tests'!F{start_row}:F{end_row}"

    formulas = [
        label,
        f"={end_row - start_row + 1}",
        f"=COUNTIF({rng},\"PASS\")",
        f"=COUNTIF({rng},\"FAIL\")",
        f"=COUNTIF({rng},\"BLOCKED\")",
        f"=IFERROR(D{row}/C{row},0)",
    ]
    for ci, (col, val) in enumerate(zip(hdr_cols, formulas)):
        c = ws_sum[f"{col}{row}"]
        c.value     = val
        c.fill      = fill(bg)
        c.font      = Font(color=OFF_W if ci > 0 else GOLD_L, size=10)
        c.alignment = center() if ci > 0 else left()
        c.border    = border()
        if ci == 5:  # Pass Rate — percentage
            c.number_format = "0%"
            c.font = Font(color=GREEN, bold=True, size=10)

# Total row
total_row = 8 + len(area_labels)
ws_sum.row_dimensions[total_row].height = 26
totals = [
    "TOTAL",
    f"=SUM(C8:C{total_row-1})",
    f"=SUM(D8:D{total_row-1})",
    f"=SUM(E8:E{total_row-1})",
    f"=SUM(F8:F{total_row-1})",
    f"=IFERROR(D{total_row}/C{total_row},0)",
]
for col, val in zip(hdr_cols, totals):
    c = ws_sum[f"{col}{total_row}"]
    c.value         = val
    c.fill          = fill(GOLD)
    c.font          = Font(color=DARK, bold=True, size=11)
    c.alignment     = center()
    c.border        = border()
    c.number_format = "0%" if col == "G" else "General"

# Overall pass rate callout
ws_sum.row_dimensions[total_row + 2].height = 40
ws_sum.merge_cells(f"B{total_row+2}:G{total_row+2}")
pr = ws_sum.cell(total_row + 2, 2)
pr.value         = f"=IFERROR(D{total_row}/C{total_row},\"—\")"
pr.number_format = "0.0%"
pr.font          = Font(color=GREEN, bold=True, size=24)
pr.fill          = fill(DARK2)
pr.alignment     = center()
pr.border        = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="medium", color=GOLD),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

lbl_row = total_row + 3
ws_sum.merge_cells(f"B{lbl_row}:G{lbl_row}")
lbl = ws_sum.cell(lbl_row, 2, "Overall Pass Rate")
lbl.font      = Font(color=GOLD_L, size=10, italic=True)
lbl.alignment = center()

# ── Save ─────────────────────────────────
out = r"C:\Users\Moham\Merchant_Club_SA_UAT.xlsx"
wb.save(out)
print(f"Saved: {out}")
