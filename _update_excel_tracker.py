"""
Updates Merchant Club SA — Execution Tracker.xlsx in-place to match
the markdown tracker (version 4 / 2026-04-28).

Changes:
  Summary sheet  — last_update header, Phase 1 done/total counts
  Phase 1 sheet  — 9 status flips to Done, 1G.6 to In Progress,
                   1C.7 task text, 1F.8 task text, append 1X section
"""
import copy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = (
    r"C:\Users\Moham\Nexus\02 Projects\Active\Merchant Club SA"
    r"\Merchant Club SA — Execution Tracker.xlsx"
)

DONE        = "✅ Done"
NOT_STARTED = "⏳ Not Started"
IN_PROGRESS = "\U0001f504 In Progress"

# ── emoji shortcuts ──────────────────────────────────────────────────────────
TICK  = "✅"
HOURGLASS = "⏳"
SPIN  = "\U0001f504"
CROSS = "\U0001f6ab"

wb = openpyxl.load_workbook(PATH)

# ════════════════════════════════════════════════════════════════════════════
# Helper — copy cell style from a reference cell
# ════════════════════════════════════════════════════════════════════════════
def copy_style(src, dst):
    dst.font      = copy.copy(src.font)
    dst.fill      = copy.copy(src.fill)
    dst.border    = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format

def set_value(ws, row, col, value):
    ws.cell(row=row, column=col).value = value

def set_status(ws, row, status, ref_row=None):
    """Set status cell (col 5) and optionally copy style from ref_row."""
    cell = ws.cell(row=row, column=5)
    cell.value = status
    if ref_row:
        copy_style(ws.cell(row=ref_row, column=5), cell)

# ════════════════════════════════════════════════════════════════════════════
# 1. SUMMARY SHEET
# ════════════════════════════════════════════════════════════════════════════
ws_sum = wb["Summary"]

# Row 2 — header bar
ws_sum.cell(row=2, column=1).value = (
    "Last updated: 2026-04-28  |  Progress: 49%  |  "
    "Phase 1 in progress  |  Target launch: 2026-08-01"
)

# Phase 1 row = row 6 — update Done (col 6) and Total (col 7)
# Old: Done=17, Total=51  →  New: Done=32 (17+9 flipped+6 new 1X), Total=57 (51+6)
ws_sum.cell(row=6, column=6).value = "32"
ws_sum.cell(row=6, column=7).value = "57"

# ════════════════════════════════════════════════════════════════════════════
# 2. PHASE 1 SHEET — status updates
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb["PHASE 1 — Model Definition & Te"]

# Reference rows for style copying
DONE_REF      = 7   # 1A.1 — existing Done row
INPROG_REF    = 11  # 1A.5 — existing In Progress row

# ── 1C section (rows 25-32) ──────────────────────────────────────────────
# 1C.2 → Done
set_status(ws1, 25, DONE, DONE_REF)
ws1.cell(row=25, column=7).value = "2026-04-19"

# 1C.3 → Done
set_status(ws1, 26, DONE, DONE_REF)
ws1.cell(row=26, column=7).value = "2026-04-19"

# 1C.4 → Done
set_status(ws1, 27, DONE, DONE_REF)
ws1.cell(row=27, column=7).value = "2026-04-19"

# 1C.5 → Done
set_status(ws1, 28, DONE, DONE_REF)
ws1.cell(row=28, column=7).value = "2026-04-19"

# 1C.7 → Done + update task description (Supabase Storage, not Vercel Blob)
ws1.cell(row=30, column=2).value = (
    "File storage (Supabase Storage) for brand logos / banners / product images — live"
)
set_status(ws1, 30, DONE, DONE_REF)
ws1.cell(row=30, column=7).value = "2026-04-19"

# ── 1F section (rows 54-61) ──────────────────────────────────────────────
# 1F.5 → Done
set_status(ws1, 58, DONE, DONE_REF)
ws1.cell(row=58, column=7).value = "2026-04-27"

# 1F.6 → Done
set_status(ws1, 59, DONE, DONE_REF)
ws1.cell(row=59, column=7).value = "2026-04-27"

# 1F.8 → Done + update task description (customer email also built)
ws1.cell(row=61, column=2).value = (
    "New order email to brand + customer confirmation email — both via Resend, fire-and-forget"
)
set_status(ws1, 61, DONE, DONE_REF)
ws1.cell(row=61, column=7).value = "2026-04-27"

# ── 1G section (rows 65-70) ──────────────────────────────────────────────
# 1G.4 → Done
set_status(ws1, 68, DONE, DONE_REF)
ws1.cell(row=68, column=7).value = "2026-04-19"

# 1G.6 → In Progress (UAT running)
set_status(ws1, 70, IN_PROGRESS, INPROG_REF)

# ════════════════════════════════════════════════════════════════════════════
# 3. PHASE 1 SHEET — append 1X section
# ════════════════════════════════════════════════════════════════════════════

# Current last row is 78 (1H.5). We add:
# Row 79 — blank
# Row 80 — section header "1X — Build Sprint Additions"
# Row 81 — column headers
# Rows 82-87 — 6 tasks

# Grab reference styles from existing section header and task rows
SECTION_HDR_REF = 72   # "1H — Brand Acquisition"
COL_HDR_REF     = 73   # '#', 'Task', 'Owner', ...
TASK_REF        = 74   # 1H.1 — a standard task row

def append_section_header(ws, row, text):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    ref = ws.cell(row=SECTION_HDR_REF, column=1)
    copy_style(ref, cell)
    # Bold the section header
    cell.font = Font(
        bold=True,
        size=ref.font.size or 10,
        color=ref.font.color.rgb if ref.font.color and ref.font.color.type == "rgb" else "000000",
        name=ref.font.name or "Calibri",
    )

def append_col_header(ws, row):
    headers = ["#", "Task", "Owner", "Priority", "Status", "Dependency", "Deadline"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = h
        copy_style(ws.cell(row=COL_HDR_REF, column=col), cell)

def append_task_row(ws, row, id_, task, owner, priority, status, dep, deadline):
    values = [id_, task, owner, priority, status, dep, deadline]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        copy_style(ws.cell(row=TASK_REF, column=col), cell)
    # Override status cell style to match Done
    if status == DONE:
        copy_style(ws.cell(row=DONE_REF, column=5), ws.cell(row=row, column=5))

# Blank separator
ws1.cell(row=79, column=1).value = None

# Section header
append_section_header(
    ws1, 80,
    "1X — Build Sprint Additions (out-of-tracker deliverables, 2026-04-27/28)"
)

# Column headers
append_col_header(ws1, 81)

# 6 task rows
X_TASKS = [
    ("1X.1", "Order status-change emails to customer (confirmed/shipped/delivered/cancelled)",
     "Abo Saif", "High", DONE, "—", "2026-04-27"),
    ("1X.2", "Auto stock decrement on order + out_of_stock flip at zero",
     "Abo Saif", "High", DONE, "—", "2026-04-27"),
    ("1X.3", "Creator member flow + admin Members page (apply, approve, reject, revoke)",
     "Abo Saif", "High", DONE, "—", "2026-04-27"),
    ("1X.4", "Dynamic sitemap — queries live brands + products, emits EN + AR URLs",
     "Abo Saif", "High", DONE, "—", "2026-04-28"),
    ("1X.5", "Locale-aware metadata — generateMetadata() returns AR or EN title/description + OG locale",
     "Abo Saif", "High", DONE, "—", "2026-04-28"),
    ("1X.6", "Custom 404 page — branded, bilingual, RTL-aware via useParams()",
     "Abo Saif", "Medium", DONE, "—", "2026-04-28"),
]

for i, task_data in enumerate(X_TASKS):
    append_task_row(ws1, 82 + i, *task_data)

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
wb.save(PATH)
print(f"Saved: {PATH}")
print("Changes applied:")
print("  Summary: last_update=2026-04-28, Phase 1 Done=32, Total=57, Progress=49%")
print("  Phase 1: 1C.2-1C.5, 1C.7, 1F.5, 1F.6, 1F.8, 1G.4 -> Done")
print("  Phase 1: 1G.6 -> In Progress")
print("  Phase 1: 1X section appended (rows 80-87, 6 tasks)")
